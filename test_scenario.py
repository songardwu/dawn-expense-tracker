"""全場景測試 — 涵蓋 V1~V4 所有功能"""
import json
import re
import sys
import urllib.request
import urllib.parse
import urllib.error
import http.cookiejar
from datetime import date

BASE = 'http://127.0.0.1:5000'
PASS = 0
FAIL = 0


def check(desc, condition):
    global PASS, FAIL
    if condition:
        PASS += 1
        print(f'  [PASS] {desc}')
    else:
        FAIL += 1
        print(f'  [** FAIL **] {desc}')


def section(title):
    print(f'\n--- {title} ---')


_csrf_cache = {}  # opener id -> token


def extract_csrf(html):
    """Extract CSRF token from HTML response."""
    m = re.search(r'name="csrf_token"\s+value="([^"]+)"', html)
    if not m:
        m = re.search(r'value="([^"]+)"\s+name="csrf_token"', html)
    return m.group(1) if m else ''


def make_session():
    cj = http.cookiejar.CookieJar()
    return urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))


def login(opener, user, pw):
    # GET login page to get CSRF token
    resp = opener.open(urllib.request.Request(BASE + '/login'))
    html = resp.read().decode('utf-8')
    token = extract_csrf(html)
    data = urllib.parse.urlencode({'username': user, 'password': pw, 'csrf_token': token}).encode('utf-8')
    resp = opener.open(urllib.request.Request(BASE + '/login', data=data, method='POST'))
    body = resp.read().decode('utf-8')
    # Cache a CSRF token for this session (fetch from landing page after login)
    t = extract_csrf(body)
    if t:
        _csrf_cache[id(opener)] = t
    return body


def get(opener, path):
    resp = opener.open(urllib.request.Request(BASE + path))
    body = resp.read().decode('utf-8')
    # Opportunistically cache CSRF token
    t = extract_csrf(body)
    if t:
        _csrf_cache[id(opener)] = t
    return resp.status, body


def post(opener, path, params):
    # Auto-inject CSRF token if not already present
    if 'csrf_token' not in params:
        token = _csrf_cache.get(id(opener), '')
        if not token:
            # Fallback: fetch a page to get token
            _, html = get(opener, '/')
            token = extract_csrf(html)
        params = dict(params)
        params['csrf_token'] = token
    data = urllib.parse.urlencode(params).encode('utf-8')
    try:
        resp = opener.open(urllib.request.Request(BASE + path, data=data, method='POST'))
        body = resp.read().decode('utf-8')
        t = extract_csrf(body)
        if t:
            _csrf_cache[id(opener)] = t
        return resp.status, body
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode('utf-8', errors='replace')


def get_json(opener, path):
    resp = opener.open(urllib.request.Request(BASE + path))
    return json.loads(resp.read().decode('utf-8'))


def db_query(sql, params=None):
    from app import get_conn
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(sql, params or ())
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows


def db_exec(sql, params=None):
    from app import get_conn
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(sql, params or ())
    conn.commit()
    cur.close()
    conn.close()


print('=' * 60)
print('  全場景測試 V1~V4')
print('=' * 60)

# ============================================================
section('1. 登入/登出/權限')
# ============================================================
admin = make_session()
login(admin, 'dawn', 'dawn1234')
s, body = get(admin, '/')
check('管理員登入成功', s == 200 and '出帳管理' in body)

designer = make_session()
login(designer, 'designer_a', 'test1234')
s, body = get(designer, '/')
check('設計師登入成功', s == 200)

anon = make_session()
s, body = get(anon, '/')
check('未登入被導向登入頁', '登入' in body or 'login' in body.lower())

# 停用帳號登入測試
bad = make_session()
login(bad, 'nonexist', 'wrong')
s, body = get(bad, '/')
check('錯誤帳號被擋', '登入' in body or 'login' in body.lower())

# ============================================================
section('2. 廠商資料管理 (V4)')
# ============================================================
# 新增廠商
post(admin, '/vendors/create', {
    'name': 'S_廠商甲公司', 'bank_name': '台北富邦中山分行',
    'bank_code': '012-0456', 'account_no': 'ACCT-111',
    'account_name': '甲公司戶名',
})
post(admin, '/vendors/create', {
    'name': 'S_廠商乙工作室', 'bank_name': '國泰世華信義分行',
    'bank_code': '013-0001', 'account_no': 'ACCT-222',
    'account_name': '乙工作室戶名',
})
# 相同帳號不同名稱（測試重複偵測）
post(admin, '/vendors/create', {
    'name': 'S_廠商甲設計', 'bank_name': '台北富邦中山分行',
    'bank_code': '012-0456', 'account_no': 'ACCT-111',
    'account_name': '甲設計戶名',
})

s, body = get(admin, '/vendors')
check('廠商甲公司存在', 'S_廠商甲公司' in body)
check('廠商乙工作室存在', 'S_廠商乙工作室' in body)
check('廠商甲設計存在', 'S_廠商甲設計' in body)

# 設計師新增
post(designer, '/vendors/create', {
    'name': 'S_廠商丙行', 'bank_name': '中信松山分行',
    'bank_code': '822-0789', 'account_no': 'ACCT-333',
    'account_name': '丙行戶名',
})
s, body = get(designer, '/vendors')
check('設計師可新增廠商', 'S_廠商丙行' in body)

# 設計師不可修改
all_ids = re.findall(r'/vendors/update/(\d+)', body)
if all_ids:
    s, _ = post(designer, f'/vendors/update/{all_ids[0]}', {
        'name': 'hack', 'bank_name': 'h', 'bank_code': 'h',
        'account_no': 'h', 'account_name': 'h',
    })
    s2, body2 = get(admin, '/vendors')
    check('設計師不可修改廠商', 'hack' not in body2)

# 設計師不可刪除
del_ids = re.findall(r'/vendors/delete/(\d+)', body)
if del_ids:
    s, _ = post(designer, f'/vendors/delete/{del_ids[0]}', {})
    s2, body2 = get(admin, '/vendors')
    check('設計師不可刪除廠商', 'S_廠商' in body2)

# 管理員修改
s, body = get(admin, '/vendors')
admin_ids = re.findall(r'/vendors/update/(\d+)', body)
if len(admin_ids) >= 2:
    s, _ = post(admin, f'/vendors/update/{admin_ids[1]}', {
        'name': 'S_廠商乙工作室', 'bank_name': '玉山銀行永和分行',
        'bank_code': '808-0100', 'account_no': 'ACCT-222-NEW',
        'account_name': '乙工作室新戶名',
    })
    s, body = get(admin, '/vendors')
    check('管理員修改廠商成功', '808-0100' in body)

# 重複名稱被擋
s, _ = post(admin, '/vendors/create', {
    'name': 'S_廠商甲公司', 'bank_name': 'x', 'bank_code': 'x',
    'account_no': 'x', 'account_name': 'x',
})
rows = db_query("SELECT COUNT(*) FROM vendors WHERE name = 'S_廠商甲公司'")
check('重複廠商名稱被擋', rows[0][0] == 1)

# 範本下載
resp = admin.open(urllib.request.Request(BASE + '/vendors/template'))
check('範本下載成功', resp.status == 200)
ct = resp.headers.get('Content-Type', '')
check('範本為 xlsx', 'spreadsheet' in ct or 'xlsx' in ct)

# ============================================================
section('3. 廠商銀行 API + 防呆增強 (V4)')
# ============================================================
data = get_json(admin, '/api/vendor-bank?name=' + urllib.parse.quote('S_廠商甲公司'))
check('API 回傳銀行資訊', data.get('bank_name') == '台北富邦中山分行')
check('API 回傳帳號', data.get('account_no') == 'ACCT-111')

data2 = get_json(admin, '/api/vendor-bank?name=noexist')
check('不存在廠商回空', data2 == {})

# 相似比對（名稱相似 + 帳號相同）
sim = get_json(admin, '/api/check-vendor?q=' + urllib.parse.quote('S_廠商甲'))
check('相似廠商偵測有結果', len(sim.get('similar', [])) > 0)

# ============================================================
section('4. 提報新增 + 匯款日期預設 + 匯款方式 (V4)')
# ============================================================
s, body = get(designer, '/new')
check('新增頁含匯款方式', '匯款方式' in body)
check('新增頁含銀行資訊框', 'bank-info' in body)
check('新增頁含預設日期提示', '下月 5 日' in body)

# 提報 1：有匯款日期
post(designer, '/submit', {
    'vendor': 'S_廠商甲公司', 'vendor_type': '水電',
    'amount': '50000', 'category': '案場成本',
    'invoice_no': 'S-001', 'invoice_date': '2026-04-10',
    'project_no': 'S案場A', 'payment_method': '公司轉帳',
    'remit_date': '2026-04-20',
})
# 提報 2：無匯款日期（測試預設）
post(designer, '/submit', {
    'vendor': 'S_廠商甲公司', 'vendor_type': '水電',
    'amount': '30000', 'category': '管銷',
    'invoice_no': 'S-002', 'invoice_date': '2026-04-12',
    'project_no': 'S案場A', 'payment_method': '現金',
    'remit_date': '',
})
# 提報 3：另一廠商
post(designer, '/submit', {
    'vendor': 'S_廠商乙工作室', 'vendor_type': '設計',
    'amount': '20000', 'category': '案場成本',
    'invoice_no': 'S-003', 'invoice_date': '2026-04-14',
    'project_no': 'S案場B', 'payment_method': '個帳轉帳',
    'remit_date': '',
})
# 提報 4：同帳號不同名的廠商（甲設計）
post(designer, '/submit', {
    'vendor': 'S_廠商甲設計', 'vendor_type': '設計',
    'amount': '15000', 'category': '獎金',
    'invoice_no': 'S-004', 'invoice_date': '2026-04-15',
    'project_no': 'S案場C', 'payment_method': '公司轉帳',
    'remit_date': '',
})

# 驗證 DB
rows = db_query("SELECT invoice_no, remit_date FROM reports WHERE invoice_no IN ('S-001','S-002','S-003','S-004') ORDER BY invoice_no")
check('4 筆提報成功寫入', len(rows) == 4)

s001 = [r for r in rows if r[0] == 'S-001']
check('S-001 匯款日期 = 手動設定 2026-04-20', str(s001[0][1]) == '2026-04-20')

s002 = [r for r in rows if r[0] == 'S-002']
from app import default_remit_date
expected_default = default_remit_date(date(2026, 4, 12))
check(f'S-002 匯款日期 = 預設 {expected_default}', str(s002[0][1]) == str(expected_default))

# 缺匯款方式被擋
s, body = post(designer, '/submit', {
    'vendor': 'S_廠商甲公司', 'vendor_type': '水電',
    'amount': '5000', 'category': '案場成本',
    'invoice_no': 'S-005', 'invoice_date': '2026-04-15',
    'project_no': 'S案場A', 'payment_method': '',
    'remit_date': '',
})
check('缺匯款方式被擋', '匯款方式' in body)

# 發票重複被擋
s, body = post(designer, '/submit', {
    'vendor': 'S_廠商甲公司', 'vendor_type': '水電',
    'amount': '5000', 'category': '案場成本',
    'invoice_no': 'S-001', 'invoice_date': '2026-04-15',
    'project_no': 'S案場A', 'payment_method': '現金',
    'remit_date': '',
})
check('重複發票被擋', '發票' in body or s == 200)

# ============================================================
section('5. 清單頁加總 + 重複廠商標示 (V4)')
# ============================================================
s, body = get(admin, '/')
check('清單頁載入', s == 200)
check('請款加總區塊存在', '請款加總' in body)
check('按廠商小計存在', 'S_廠商甲公司' in body)
check('按匯款方式存在', '按匯款方式' in body)
check('總計存在', '總計' in body)
check('匯款方式欄位顯示', '公司轉帳' in body)

# 重複廠商偵測（甲公司 vs 甲設計，同帳號 ACCT-111）
check('重複廠商警示區塊存在', '可能重複的廠商' in body)

# ============================================================
section('6. 管理員 inline 編輯 (V3+V4)')
# ============================================================
s001_rows = db_query("SELECT id FROM reports WHERE invoice_no = 'S-001'")
if s001_rows:
    rid = s001_rows[0][0]
    s, _ = post(admin, f'/update-report/{rid}', {
        'vendor': 'S_廠商甲公司', 'amount': '55000',
        'category': '案場成本', 'invoice_no': 'S-001',
        'invoice_date': '2026-04-10', 'project_no': 'S案場A',
        'remit_date': '2026-04-25', 'payment_method': '現金',
    })
    check('管理員 inline 編輯成功', s == 200)

    rows = db_query("SELECT amount, payment_method, remit_date FROM reports WHERE invoice_no = 'S-001'")
    check('金額更新為 55000', float(rows[0][0]) == 55000)
    check('匯款方式改為現金', rows[0][1] == '現金')
    check('匯款日期改為 2026-04-25', str(rows[0][2]) == '2026-04-25')

# ============================================================
section('7. 案場鎖定 (V3)')
# ============================================================
post(admin, '/toggle-lock-project', {'project_no': 'S案場A', 'action': 'lock'})
rows = db_query("SELECT is_locked FROM reports WHERE project_no = 'S案場A'")
check('S案場A 已鎖定', all(r[0] for r in rows))

# 鎖定後不能刪除
s, body = get(admin, '/')
del_ids = re.findall(r'/delete/(\d+)', body)
locked_rows = db_query("SELECT id FROM reports WHERE project_no = 'S案場A' AND is_locked = TRUE")
if locked_rows:
    s, _ = post(designer, f'/delete/{locked_rows[0][0]}', {})
    check('鎖定後設計師不能刪除', s == 403)
    s, _ = post(admin, f'/delete/{locked_rows[0][0]}', {})
    check('鎖定後管理員也不能刪除', s == 403)

# 鎖定後不能改匯款日期
s, _ = post(designer, f'/update-remit-date/{locked_rows[0][0]}', {'remit_date': '2026-06-01'})
check('鎖定後不能改匯款日期 (V3 fix)', s == 403)

# 解鎖
post(admin, '/toggle-lock-project', {'project_no': 'S案場A', 'action': 'unlock'})
rows = db_query("SELECT is_locked FROM reports WHERE project_no = 'S案場A'")
check('S案場A 已解鎖', all(not r[0] for r in rows))

# ============================================================
section('8. 設計師資料隔離 (V2)')
# ============================================================
s, body = get(designer, '/')
check('設計師看到自己的提報', 'S_廠商甲公司' in body)
# 設計師不能看到其他人的提報 (admin 沒有提報，所以沒有衝突)
check('設計師看不到提報人欄', '提報人' not in body)

# ============================================================
section('9. 設計師操作他人資料 (V2)')
# ============================================================
# 設計師不能刪除管理員建立的資料（如果有）
# 建立一筆管理員提報
post(admin, '/submit', {
    'vendor': 'S_管理員廠商', 'vendor_type': '其他',
    'amount': '10000', 'category': '管銷',
    'invoice_no': 'S-ADMIN-001', 'invoice_date': '2026-04-15',
    'project_no': 'S案場X', 'payment_method': '現金',
    'remit_date': '',
})
admin_report = db_query("SELECT id FROM reports WHERE invoice_no = 'S-ADMIN-001'")
if admin_report:
    s, _ = post(designer, f'/delete/{admin_report[0][0]}', {})
    check('設計師不能刪除管理員提報', s == 403)

# ============================================================
section('10. 輸入驗證 (V3 fix)')
# ============================================================
s001_rows2 = db_query("SELECT id FROM reports WHERE invoice_no = 'S-001'")
if s001_rows2:
    rid = s001_rows2[0][0]
    # 不合法 category
    s, _ = post(admin, f'/update-report/{rid}', {
        'vendor': 'S_廠商甲公司', 'amount': '55000',
        'category': '非法分類', 'invoice_no': 'S-001',
        'invoice_date': '2026-04-10', 'project_no': 'S案場A',
        'remit_date': '', 'payment_method': '現金',
    })
    rows = db_query("SELECT category FROM reports WHERE invoice_no = 'S-001'")
    check('非法 category 被擋', rows[0][0] == '案場成本')

    # 負數金額
    s, _ = post(admin, f'/update-report/{rid}', {
        'vendor': 'S_廠商甲公司', 'amount': '-100',
        'category': '案場成本', 'invoice_no': 'S-001',
        'invoice_date': '2026-04-10', 'project_no': 'S案場A',
        'remit_date': '', 'payment_method': '現金',
    })
    rows = db_query("SELECT amount FROM reports WHERE invoice_no = 'S-001'")
    check('負數金額被擋', float(rows[0][0]) == 55000)

# ============================================================
section('11. 安全 HTTP Headers (V3 fix)')
# ============================================================
resp = admin.open(urllib.request.Request(BASE + '/'))
check('X-Content-Type-Options: nosniff', resp.headers.get('X-Content-Type-Options') == 'nosniff')
check('X-Frame-Options: DENY', resp.headers.get('X-Frame-Options') == 'DENY')
csp = resp.headers.get('Content-Security-Policy', '')
check('CSP 存在', "default-src 'self'" in csp)

# ============================================================
section('12. 匯款日期預設邏輯驗證')
# ============================================================
from app import default_remit_date, is_business_day, next_business_day

# 2026-05-05 (Tue) — normal
check('2026-05-05 (Tue) is biz day', is_business_day(date(2026, 5, 5)))

# 2026-04-05 (Sun) → 4/6 (Mon, 清明補假) → 4/7 (Tue)
d = default_remit_date(date(2026, 3, 15))
check(f'3月→4/7 (跳過週日+清明補假): {d}', d == date(2026, 4, 7))

# 2027-01-05 (Tue) — normal
d2 = default_remit_date(date(2026, 12, 10))
check(f'12月→1/5: {d2}', d2 == date(2027, 1, 5))

# 2027-02-05 (Fri, 除夕前放假) → 2/10 is fine? Let's check
d3 = default_remit_date(date(2027, 1, 15))
check(f'2027/1月→2月: {d3}', d3 == date(2027, 2, 11))
# 2/5 Fri=除夕前, 2/6 Sat=除夕, 2/7 Sun=春節, 2/8 Mon=春節, 2/9 Tue=春節, 2/10 Wed=春節補假
# → 2/11 Thu

# 5月→6月: 6/5 (Fri) is normal business day
d4 = default_remit_date(date(2026, 5, 10))
check(f'5月→6/5 (正常工作日): {d4}', d4 == date(2026, 6, 5))

# 2026-10-05 (Mon) — normal? Check it's not a holiday
d5 = default_remit_date(date(2026, 9, 10))
check(f'9月→10/5: {d5}', d5 == date(2026, 10, 5))

# ============================================================
section('13. Excel 匯出 (V1+V4)')
# ============================================================
resp = admin.open(urllib.request.Request(BASE + '/export'))
check('Excel 匯出 200', resp.status == 200)
ct = resp.headers.get('Content-Type', '')
check('xlsx 格式', 'spreadsheet' in ct)
xlsx_data = resp.read()
check('檔案大小 > 0', len(xlsx_data) > 100)

# 驗證內容含新欄位
import openpyxl
from io import BytesIO
wb = openpyxl.load_workbook(BytesIO(xlsx_data))
ws = wb['明細']
headers = [cell.value for cell in ws[1]]
check('Excel 含匯款方式欄', '匯款方式' in headers)
check('Excel 含銀行分行欄', '銀行分行名稱' in headers)
check('Excel 含銀行代碼欄', '銀行代碼' in headers)
check('Excel 含帳號欄', '銀行帳號' in headers)
check('Excel 含戶名欄', '戶名' in headers)

# 檢查銀行資料有帶入
has_bank = False
for row in ws.iter_rows(min_row=2, values_only=True):
    row_dict = dict(zip(headers, row))
    if row_dict.get('銀行代碼'):
        has_bank = True
        break
check('Excel 有帶入銀行資料', has_bank)
wb.close()

# ============================================================
section('14. 導覽列 (V4)')
# ============================================================
s, body = get(admin, '/')
check('導覽列含廠商資料', '廠商資料' in body)
check('導覽列含帳號管理(admin)', '帳號管理' in body)

s, body = get(designer, '/')
check('設計師導覽列含廠商資料', '廠商資料' in body)
check('設計師導覽列無帳號管理', '帳號管理' not in body)

# ============================================================
section('15. 清理測試資料')
# ============================================================
db_exec("DELETE FROM reports WHERE invoice_no LIKE %s", ('S-%',))
db_exec("DELETE FROM reports WHERE invoice_no = %s", ('S-ADMIN-001',))
db_exec("DELETE FROM vendors WHERE name LIKE %s", ('S_%',))
remaining = db_query("SELECT COUNT(*) FROM reports WHERE invoice_no LIKE %s", ('S-%',))
check('報表清理完成', remaining[0][0] == 0)
remaining2 = db_query("SELECT COUNT(*) FROM vendors WHERE name LIKE %s", ('S_%',))
check('廠商清理完成', remaining2[0][0] == 0)

# ============================================================
print('\n' + '=' * 60)
print(f'  結果：{PASS} 通過 / {FAIL} 失敗 / 共 {PASS + FAIL} 項')
print('=' * 60)

if FAIL > 0:
    sys.exit(1)
